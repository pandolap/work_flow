#!/usr/bin/env python
# -*- coding: UTF-8 -*-

import json
import requests
import re, os
from requests_toolbelt import MultipartEncoder
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, timedelta


def imageJson(url, count=5):
    if count < 0:
        return
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/"
                      "537.36 (KHTML, like Gecko) Chrome/75.0.3770.80 Saf"
                      "ari/537.36",
    }
    image_content = requests.get(url, headers=headers).content

    m = MultipartEncoder(
        fields={
            'file': ('images.jpg', image_content, 'text/plain'),
            'uuid': '2563f6b39b6d42c29d28ec3eb8e1d00f',
            'checkInvoice': 'True',
            'ifOcrType': 'False',
        }
    )
    try:
        omg = requests.post(url="http://ocr.pre.ninetechone.com/fileOCRAndcheckInvoice", data=m,
                            headers={'Content-Type': m.content_type})
        # return omg.json()
        rs = omg.json()
        if rs.get('message', '').find('error') > -1:
            return imageJson(url, count=count - 1)
        if rs.get('error', 0) > 0:
            return imageJson(url, count=count - 1)
        return rs
    except:
        count = count - 1
        return imageJson(url, count)


def keep_log(log_dir, keep_day):
    for fn in os.listdir(log_dir):
        if fn.find('_') < 0:
            continue
        dt = fn.split('_')[1].strip('.txt')
        t = datetime.strptime(dt, '%Y-%m-%d')
        now = datetime.now()
        if now - t > timedelta(days=keep_day):
            try:
                os.remove(os.path.join(log_dir, fn))
            except:
                pass

data = json.loads(flow)
#keep_log(data['config']['log_dir'], data['config']['keep_day'])
now_date = datetime.now().strftime('%Y-%m-%d')
log_file_name = os.path.join(data['config']['log_dir'], '应付日志_%s.txt' % now_date)
baseInfo = data['data']['baseInfo']
home_dir = data['config']['home_dir']

def log(s):
    with open(log_file_name, 'a', encoding='utf-8') as f:
        try:
            f.write(str(s))
        except:
            pass


# Todo url_dict_new 是所有图片的链接 return result.images.N
url_dict_new = data['data'].get('imgN', [])
# 记录url地址

log('*' * 60)
log('\n')
log('单据编号: ')
log(baseInfo.get('单据编号'))
log('\n')
for line in url_dict_new:
    log(line)
    log('\n')

# Todo advice 是之前所有的意见
# advice = data['data'].get('verifyResult', '')+'\n'
advice = ''

# Todo costBearingcompany 是费用承担公司 深圳妈港仓码有限公司
costBearingcompany = baseInfo.get('公司')

# Todo receivingBank 是收款银行 中国工商银行股份有限公司深圳梅林支行
receivingBank = baseInfo.get('收款银行')

# Todo payee 是收款人  深圳柏汇自动化设备有限公司
payee = baseInfo.get('收款人').replace('(', '').replace(')', '').replace('（', '').replace('）', '').replace(' ', '')

# Todo collectionAccount 收款账户   4000026209200366760
collectionAccount = baseInfo.get('收款账户')

# Todo moneyFee 合计税额   106.99
moneyFee = baseInfo.get('合计税额')
moneyFee = moneyFee.replace(',', '')
if moneyFee == ' ' or moneyFee == '':
    moneyFee = baseInfo.get('税率')
    moneyFee = moneyFee.replace(',', '')
# Todo feeSum 应付金额      0
feeSum = baseInfo.get('应付金额')
feeSum = feeSum.replace(',', '')

totalSum = 0
taxSum = 0
listExcel = []
advice_first = advice

bearing_flag = None  # 费用承担公司标识
taxnumber_flag = None  # 税号标识
seal_flag = None  # 盖章标识
formname_flag = None  # 发票联标识
receivebank_flag = None  # 收款银行标识
seller_flag = None  # 销售方标识
collectionAccount_flag = None  # 收款账户标识
flight_flag = None  # 机票仓位标识
seat_flag = None  # 货车仓位标识
question_list = []
checkinvoice_flag = None  # 发票验真标识
# try:
# for i in url_dict['images']['N']:
for url_index, url in enumerate(url_dict_new):
    # url = i['url']  # 拿到每一条票据的url地址
    url_index += 1
    try:
        omg = imageJson(url)
        try:
            if omg["response"]:
                for item in omg['response']['data']['identify_results']:
                    type = item['type']
                    try:
                        if type == '1001' and item['details']['form_name'] == '抵扣联':
                            continue
                    except:
                        pass
                    if type == '0000':
                        # advice = advice + '没有检测到发票；'
                        pass
                    if type == '1001' or type == '1002' or type == '1003' or type == '1004':  # 增值税普通发票
                        # 购买方名称
                        buyer = item['details']['buyer']
                        buyer = buyer.replace('（', '').replace('）', '').replace('(', '').replace(')', '')
                        if omg['checkInvoice'] == True:
                            checkinvoice_flag = 0
                        else:
                            checkinvoice_flag = 1
                            question_list.append(url_index)
                        # hlt添加、总分公司对应
                        sub_company_col = 3
                        father_company_col = 4
                        company_dict = {}
                        wb = openpyxl.load_workbook(os.path.join(home_dir, '港口公司费用承担公司名称核对.xlsx'))
                        table = wb.get_sheet_by_name('Sheet2')
                        max_row = table.max_row
                        for irow in range(max_row):
                            irow += 1
                            relation = {table.cell(irow, sub_company_col).value: table.cell(irow, father_company_col).value}
                            company_dict.update(relation)
                        if costBearingcompany in company_dict.keys():
                            costBearingcompany = company_dict[costBearingcompany]

                        costBearingcompany = costBearingcompany.replace('（', '').replace('）', '').replace('(', '').replace(
                            ')', '')
                        # advice = advice.replace('该单没有发票；','')
                        if buyer == costBearingcompany:
                            print('购买方名称与费用承担公司一致；')
                            bearing_flag = 0
                        else:
                            # advice = advice + '购买方名称与费用承担公司不一致；'
                            bearing_flag = 1
                            question_list.append(url_index)
                        # 纳税人识别号
                        buyer_tax_id = item['details']['buyer_tax_id']
                        excel_data = pd.read_excel(os.path.join(home_dir, '纳税人识别号.xlsx'))
                        excelSerise = pd.Series(excel_data['企业名称'])
                        excelNp = np.array(excelSerise)
                        for value in excelNp:
                            listExcel.append(value)
                        if buyer in listExcel:
                            if buyer_tax_id == excel_data['税号'][listExcel.index(buyer)]:
                                # advice = advice + '税号一致；'
                                taxnumber_flag = 0
                            else:
                                # advice = advice + '税号不一致；'
                                taxnumber_flag = 1
                                question_list.append(url_index)

                        try:
                            # 发票盖章
                            company_seal = item['details']['company_seal']
                            if company_seal == '1':
                                # advice = advice + '第{}张影像中发票有盖章；'.format(url_index)
                                seal_flag = 0
                            elif company_seal == '0':
                                # advice = advice + '第{}张影像中发票没有盖章；'.format(url_index)
                                seal_flag = 1
                                question_list.append(url_index)
                        except:
                            # advice = advice + '第{}张影像中发票没有盖章；'.format(url_index)
                            seal_flag = 1
                            question_list.append(url_index)
                        try:
                            # 发票联
                            if type == '1003':
                                pass
                            else:
                                form_name = item['details']['form_name']
                                if form_name:
                                    # advice = advice + '第{}张影像中发票有发票联；'.format(url_index)
                                    formname_flag = 0
                                else:
                                    # advice = advice + '第{}张影像中发票没有发票联；'.format(url_index)
                                    formname_flag = 1
                                    question_list.append(url_index)
                        except:
                            # advice = advice + '第{}张影像中发票没有发票联；'.format(url_index)
                            formname_flag = 1
                            question_list.append(url_index)

                    if type == '1001' or type == '1002' or type == '1003' or type == '1004':
                        # 价税合计
                        total = item['details']['total']
                        totalSum = totalSum + round(float(total), 2)
                        # 销售方名称
                        seller = item['details']['seller'].strip().replace('(', '').replace(')', '').replace('（', '').replace(
                            '）', '')
                        # 销售方开户行及账号
                        if type == '1004':
                            pass
                        else:
                            seller_bank_account = item['details']['seller_bank_account']
                            # 销售方开户行
                            try:
                                seller_bank = re.findall(r'\d+', seller_bank_account)[0].strip()
                                if ':' in seller_bank:
                                    seller_bank = seller_bank.replace(':', '')
                                # 销售方账号
                                seller_account = re.findall(r'\d+\.?\d*', seller_bank_account)[0]

                                # 银行关键字
                                bankCompany = ['中国银行', '中行', '农业银行', '农行', '工商银行', '工行', '建设银行', '建行', '招商银行', '招行']

                                if seller_bank == receivingBank:
                                    print('收款银行与发票影像中的销售方开户行一致')
                                else:
                                    for bankItem in bankCompany:
                                        if bankItem in seller_bank:
                                            if bankItem in receivingBank:
                                                receivebank_flag = 0
                                                print('收款银行与发票影像中的销售方开户行一致')
                                            else:
                                                # advice = advice + '第{}张影像中收款银行与发票影像中的销售方开户行不一致；'.format(url_index)
                                                receivebank_flag = 1
                                                question_list.append(url_index)
                            except:
                                receivebank_flag = 1
                                question_list.append(url_index)
                            if payee == seller:
                                seller_flag = 0
                                print('第{}张影像中收款人与发票中的销售方名称一致'.format(url_index))
                            else:
                                # advice = advice + '第{}张影像中收款人与发票中的销售方名称不一致；'.format(url_index)
                                seller_flag = 1
                                question_list.append(url_index)

                            # if receivingBank == seller_bank:
                            #     print('收款银行与发票影像中的销售方开户行一致')
                            # else:
                            #     advice = advice + '收款银行与发票影像中的销售方开户行不一致；'

                            if collectionAccount.strip() in seller_account.strip():
                                collectionAccount_flag = 0
                                print('收款账户与发票影像中的销售方账号一致')
                            else:
                                # advice = advice + '第{}张影像中发票收款账户与发票影像中的销售方账号不一致；'.format(url_index)
                                collectionAccount_flag = 1
                                question_list.append(url_index)

                    if type == '1001':  # 增值税专用发票
                        # 税额
                        tax = item['details']['tax']
                        if ',' in tax:
                            tax = tax.replace(',', '')
                        taxSum = taxSum + float(tax)  # 累加专项发票

                    if type == '1044':  # 机票或者行程单
                        # 仓位
                        flights = item['details']['flights']
                        for omg in flights:
                            seat = omg['seat']
                            if 'F' in seat or 'A' in seat or 'C' in seat or 'D' in seat:
                                # advice = advice + '第{}张影像中机票或行程单超仓；'.format(url_index)
                                flight_flag = 1
                                question_list.append(url_index)
                            else:
                                # advice = advice + '第{}张影像中机票或行程单正常；'.format(url_index)
                                flight_flag = 0

                    if type == '1042':  # 火车票、高铁票
                        # 仓位
                        toSeat = item['details']['seat']
                        if '一等座' in toSeat or '商务座' in toSeat:
                            # advice = advice + '第{}张影像中火车票或高铁票超仓；'.format(url_index)
                            seat_flag = 1
                            question_list.append(url_index)
                        else:
                            seat_flag = 0
        except:
            advice += '第{}张影像接口无返回值；'.format(url_index)
    except Exception as e:
        print(e)
        log('\n***************：\n')
        import datetime
        log(str(datetime.datetime.now()))
        log(str(e))
        log('\n')
        advice += '流程异常，请联系运维人员排查；'.format(url_index)

# HLT修改后逻辑
if bearing_flag == 0:
    advice += '费用承担公司与发票抬头一致；'
elif bearing_flag == 1:
    advice += '【费用承担公司与发票抬头不一致】；'
if taxnumber_flag == 0:
    advice += '税号一致；'
elif taxnumber_flag == 1:
    advice += '【税号不一致】；'
if seal_flag == 0:
    advice += '发票有盖章；'
elif seal_flag == 1:
    advice += '【发票没有有盖章】；'
if formname_flag == 0:
    advice += '发票有发票联；'
elif formname_flag == 1:
    advice += '【发票没有发票联】；'
if checkinvoice_flag == 0:
    advice += '发票验真通过；'
elif checkinvoice_flag == 1:
    advice += '【发票验真不通过】；'
# 核对明细中的税额与专票增值税额
try:
    moneyFee = float(moneyFee)
except:
    moneyFee = 0
try:
    feeSum = float(feeSum)
except:
    feeSum = 0
if round(float(taxSum), 2) == float(moneyFee):
    print('专票税额正常')
elif round(float(taxSum), 2) == 0:
    print('专票税额正常')
else:
    advice = advice + '【专票税额待核查】；'

# 核对明细中的应付金额与所有发票的价税合计
if float(feeSum) == round(float(totalSum), 2):
    # float(totalSum)
    print('应付金额正常')
elif round(float(totalSum), 2) == 0:
    print('应付金额正常')
else:
    advice = advice + '【应付金额待核查】；'

if bearing_flag == taxnumber_flag == seal_flag == formname_flag == receivebank_flag == seller_flag == collectionAccount_flag == flight_flag == seat_flag == None:
    pass
else:
    taxSum = round(float(taxSum), 2)
    totalSum = round(float(totalSum) * 100) / 100.0
    advice = advice + '发票金额：' + str(totalSum) + '；' + '专票税额：' + str(taxSum) + '；'

if receivebank_flag == 0:
    advice += '\n收款银行一致；'
elif receivebank_flag == 1:
    advice += '\n【收款银行不一致】；'
if seller_flag == 0:
    advice += '收款人一致；'
elif seller_flag == 1:
    advice += '【收款人不一致】；'
if collectionAccount_flag == 0:
    advice += '收款账号一致；'
elif collectionAccount_flag == 1:
    advice += '【收款账号不一致】；'
if flight_flag == 0:
    advice += '机票或行程单正常；'
elif flight_flag == 1:
    advice += '【机票或行程单超仓】；'
if seat_flag == 0:
    advice += '火车票或高铁票正常；'
elif seat_flag == 1:
    advice += '【火车票或高铁票超仓】；'
question_list = list(set(question_list))

if '不一致' in advice or '没有' in advice or '超仓' in advice or '不通过' in advice:
    if len(question_list) == 0:
        pass
    else:
        a = ''
        for i in question_list:
            i = str(i)
            a += i + ','
        a = a[:-1]
        advice += '【请检查第{}张影像】；'.format(a)

if (
        bearing_flag == taxnumber_flag == seal_flag == formname_flag == receivebank_flag == seller_flag == collectionAccount_flag == flight_flag == seat_flag == None):
    advice = advice + '【没有检测到发票】；'

log('\nocr一次处理\n结果：\n')
log(advice)
log('\n')
print(advice)
data['data']['ocr'] = '发票信息检验：' + advice
if len(data['data'].get('nginxInfo', '')) > 0:
    nginxInfo = data['data']['nginxInfo'].strip().replace('\n', '')
    data['data']['ocr'] += '打开影像系统页面失败,提示:%s' % nginxInfo
flow = json.dumps(data)
