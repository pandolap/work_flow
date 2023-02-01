#!/usr/bin/env python
# -*- coding: UTF-8 -*-

# 统计数据计算

import os
import shutil
import zipfile
import openpyxl
import numpy as np
import pandas as pd
from barcode import Code128
from datetime import datetime
import win32com.client as win32
from barcode.writer import ImageWriter
from openpyxl.drawing.image import Image
from openpyxl.utils.units import pixels_to_EMU
# from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Font, Border, Alignment, Side
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor

# 入参
dirs = {
    'input': r'C:\\RPA\\中外运_博世叫料\\input',
    'output': r'C:\\RPA\\中外运_博世叫料\\output',
    'runtime': r'C:\\RPA\\中外运_博世叫料\\output\\20230201',
    'output_back': r'C:\\RPA\\中外运_博世叫料\\output_back',
    'runtime_back': r'C:\\RPA\\中外运_博世叫料\\output_back\\20230201',
    'log': r'C:\\RPA\\中外运_博世叫料\\output\\20230201\\log',
    'screenshot': r'C:\\RPA\\中外运_博世叫料\\output\\20230201\\err_img',
    'download': r'C:\\RPA\\中外运_博世叫料\\output\\20230201\\download',
    'supply_order': r'C:\\RPA\\中外运_博世叫料\\output\\20230201\\order'
}

# 邮箱附件地址
# attach_path = r"C:\RPA\中外运_博世叫料\output\20221215\download\邮箱附件.xlsx"
# 邮箱映射
supplier_email = {
    "DL": "aoyucs@163.com",
    "MYS": "Fupj@szmys.com",
    "YT": "53237761@qq.com",
    "WZXC": "cssales2@szwzxc.com",
}
# 文件映射
file_dict = {
    "attach": r"C:\RPA\中外运_博世叫料\output\20230201\download\20230131_SAP_97506873.xlsx",
    "export": r"C:\\RPA\\中外运_博世叫料\\output\\20230201\\download\\111平台-BSZZ01-库存即时报表（自用）.xlsx",
    "template": r"C:\\RPA\\中外运_博世叫料\\output\\20230201\\原材料入库模板.xlsx",
    "supply": r"C:\\RPA\\中外运_博世叫料\\output\\20230201\\供应商送货单.xlsx",
    "material": r"C:\\RPA\\中外运_博世叫料\\output\\20230201\\原材料需求叫料.xlsx",
    "not_receipt": r"C:\\RPA\\中外运_博世叫料\\output\\20230201\\download\\已叫料但未收货.xlsx"
}
# 缩写映射
abbr_dict = {
    "鼎立": "DL",
    "王子新材": "WZXC",
    "美盈森": "MYS",
    "裕同": "YT",
    "博世": "BS",
    "康乐": "KL",
    "沈阳防锈": "SYFX",
    "鑫潞": "XL"
}

print(supplier_email)


# 检测传入的文件是否存在
def check_file_exist(*target_files):
    if target_files:
        res = [str(file) for file in target_files if file is None or os.path.exists(file) is False]
        if res:
            return False, ",".join(res)
        else:
            return True, None
    else:
        return False, target_files


# 多文件压缩
def files_zip(to_zip_path: str, zip_name: str, files: list) -> str:
    # 获取一个
    zip_path = os.path.join(to_zip_path, "{}.zip".format(zip_name))
    # 如果存在则删除原有的
    if os.path.exists(zip_path):
        os.remove(zip_path)
    with zipfile.ZipFile(zip_path, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for f in files:
            parent_path, file_name = os.path.split(f)
            zf.write(f, arcname=file_name)
    return zip_path


# 将dataframe数据组合成二维数据
def copy_process(df_copy):
    # 处理将要拷贝的数据
    compose = [np.array(df_copy.columns)]
    # 拼接主体数据
    compose.extend(df_copy.values)
    return compose


# 对原有数据进行覆写
def form_override(data, target_file, sheet_name=None, start_row=0):
    # 检测文件的正确性
    (is_exist, files) = check_file_exist(target_file)
    if data and is_exist and sheet_name:
        if isinstance(data, list):
            wb = openpyxl.load_workbook(target_file)
            try:
                # 判断是否为默认的sheet页
                if sheet_name:
                    sheet = wb[sheet_name]
                else:
                    sheet = wb.active
                # 获取数据的行数
                row_max = len(data)
                # 进行数据覆写
                for row in range(0, row_max):
                    for col in range(0, len(data[row])):
                        # 写入一个单元格
                        sheet.cell(row=start_row + 1, column=col + 1, value=data[row][col])
                    start_row += 1
                # 将数据保存
                wb.save(target_file)
            except Exception as e:
                raise Exception("数据写入失败|<{}>".format(e))
            finally:
                # 关闭文件传输
                wb.close()
        else:
            raise TypeError("传入的覆写数据类型错误")
    else:
        raise ValueError("覆写数据未空或是目标文件错误")


# 刷新表格文件
def excel_refresh(target_excel):
    (is_exist, files) = check_file_exist(target_excel)
    # 检测文件稳定性
    if is_exist:
        app = win32.DispatchEx("Excel.Application")
        # 设置为不可见
        app.Visible = False
        # 设置为不警告
        app.DisplayAlerts = False
        try:
            wb = app.Workbooks.Open(target_excel)
            wb.RefreshAll()
            app.CalculateUntilAsyncQueriesDone()
            wb.Save()
        except Exception as e:
            raise IOError("文件刷新失败|<{}>".format(e))
        finally:
            app.Quit()
    else:
        raise FileNotFoundError("找不到将要刷新的文件|<{}>".format(target_excel))


# 系统库存导出数据写入 => 进化 = 通用拷贝表格数据方法
def copy_excel_data(from_file, to_file, to_sheet=None, from_row=0, is_cover=True, copy_area=None):
    """
    将源自A表中的n列数据拷贝到B表中
    :param from_file: 数据来源文件
    :param to_file: 数据去向文件
    :param to_sheet: 拷贝到的表格sheet名默认为第一个
    :param from_row: 数据来源从第几行开始
    :param is_cover: 拷贝是否覆盖原有数据行
    :param copy_area: 拷贝来源的范围列入："A,C,D:Q"
    """
    # 检测文件的正确性
    (is_exist, files) = check_file_exist(from_file, to_file)
    if is_exist:
        # 读取导出数据【物料代码】和【未分配】
        df = pd.read_excel(from_file, header=from_row, usecols=copy_area)
        # 首先判断是否为空
        if not len(df):
            return
        # 处理拷贝数据
        copy_data = copy_process(df)
        # 将数据写入目标文件
        if is_cover:
            form_override(copy_data, to_file, to_sheet)
        else:
            # 计算当前有多少行数据
            df = pd.read_excel(to_file, to_sheet)
            exist_row = len(df.index)
            form_override(copy_data, to_file, to_sheet, exist_row + 1)
    else:
        if files:
            raise FileNotFoundError("文件未找到|<{}>".format(files))
        else:
            raise ValueError("传入的文件值为空|<{}>".format(files))


def create_export_data(order_no, supplier_type, goods_code, receipt, stock_local):
    # 初始化单挑入库数据实例
    template_entity = {
        "货主代码": "BSZZ",
        "外部订单号": order_no,
        "订单类型": "YLRK",
        "供应商名称": supplier_type,
        "货品代码": goods_code,
        "预期收货数量": receipt,
        "包装单位": "EA",
        "库位": stock_local,
        "批次状态": "A",
        "物料状态": "N"
    }
    return template_entity


def subtotal_data(supplier_name, supplier_abbr, group_data, refer):
    # 4、获取需求订单数据（订单数量 - 半成品库存）
    # 获取当前日期字符串
    current_time = datetime.now().strftime("%Y%m%d")
    # 3、批次列表
    export_data = []
    for row in group_data.itertuples():
        # 货品代码
        goods_code = row[0][0]
        # 库位
        stock_local = row[0][3]
        if stock_local == 0:
            stock_local = "L01"

        # 半成品库存
        stock = int(row[0][1])

        # 获取分料数量
        sort_num = int(row[1])

        # 获取共用料号数量
        query_condition = "共用二级料号=='{}'".format(goods_code)
        share_num = refer.query(query_condition).get("需求数量")
        # 判断共用料号是否存在
        if len(share_num) == 1:
            share_num = int(share_num)
        else:
            share_num = 0

        # 获取原材料需求数量
        need_amount = share_num
        if share_num == 0:
            need_amount = sort_num - stock

        # 判断需求数量是否大于0
        if need_amount > 0:
            order_no = "{}{}".format(current_time, supplier_abbr)
            export_data.append(create_export_data(order_no, supplier_name, goods_code, need_amount, stock_local))

    return export_data


def load_template(file_path, data):
    """
    加载模板
    :param file_path: 模板文件
    :param data: 将要写入的数据
    """
    # 读取模板文件格式表
    df = pd.read_excel(file_path, header=1)
    # 将数据填入格式表中
    for row in data:
        df.loc[len(df)] = row
    # 导出数据为二维数组
    out_value = df.values
    # 获取导入数据量
    idx = len(out_value)
    # 打开模板文件
    wb = openpyxl.load_workbook(file_path)
    # 激活第一个sheet
    sheet = wb.active
    # 将数据写入文档
    for i in range(0, idx):
        for j in range(0, len(out_value[i])):
            cell = sheet.cell(row=2 + i + 1, column=j + 1, value=out_value[i][j])
            # 设置单元格样式
            set_cell_style_by_export(cell)
    # 保存结果并关闭链接
    wb.save(file_path)
    wb.close()


def get_barcode(code, img_path):
    if not os.path.exists(img_path):
        os.makedirs(img_path)
    encode = Code128(code, writer=ImageWriter())
    file_path = os.path.join(img_path, code)
    encode.save(file_path, options={
        'font_size': 3,
        'quiet_zone': 1,
        'module_width': 0.173,
        'module_height': 4,
        'text_distance': 1.3
    })
    # encode = Code128Encoder(code)
    # encode.save(img_path)
    return "{}.png".format(file_path)


def offset_img(img):
    """精确设置图片位置，偏移量以万为单位进行微调吧，具体计算公式太麻烦了
    row column 的索引都是从0开始的，我这里要把图片插入到单元格B10
    """
    p2e = pixels_to_EMU
    h, w = img.height, img.width
    size = XDRPositiveSize2D(p2e(w), p2e(h))
    marker = AnchorMarker(col=3, colOff=850000, row=1, rowOff=40000)
    img.anchor = OneCellAnchor(_from=marker, ext=size)


def generate_order_by_supply(order, data, supply_file, to_dir):
    data_list = []
    # 获取需要的数据
    row_idx = 0
    for row in data:
        order_no = row.get("外部订单号")
        if order_no != order:
            continue
        row_idx += 1
        entity = [row_idx]
        # 物料代码
        code = row.get("货品代码")
        entity.append(code)
        # 供应商
        supply = row.get("供应商名称")
        entity.append(supply)
        # 数量 -> 计划数量
        amount = row.get("预期收货数量")
        entity.append(amount)
        # 单位
        entity.append("PCS")
        # 上架库位
        stock_local = row.get("库位")
        entity.append(stock_local)
        # 实发数量
        entity.append("")
        # 备注
        entity.append("")
        data_list.append(entity)

    # 写入供货单
    wb = openpyxl.load_workbook(supply_file)
    sheet = wb.active

    # 单号
    sheet.cell(row=2, column=1, value=order)

    # 条码
    barcode_folder = os.path.join(to_dir, "barcode")
    barcode_path = get_barcode(order, barcode_folder)
    barcode_img = Image(barcode_path)
    # 计算条码长宽
    # img_height = sheet.row_dimensions[2].height
    # img_width = sheet.column_dimensions["E"].width + sheet.column_dimensions["F"].width + \
    #             sheet.column_dimensions["G"].width
    # 508 img_height * 3.26
    # barcode_img.height = 90
    barcode_img.width = 300
    offset_img(barcode_img)
    # img_width * 3.11
    # _from = AnchorMarker(7, 50000, 1, 50000)
    # to = AnchorMarker(7, -50000, 1, -50000)
    # barcode_img.anchor = TwoCellAnchor('twoCell', _from, to)
    sheet.add_image(barcode_img)

    idx = len(data_list)
    for i in range(0, idx):
        for j in range(0, len(data_list[i])):
            sheet.row_dimensions[3 + i + 1].height = "25"
            cell = sheet.cell(row=3 + i + 1, column=j + 1, value=data_list[i][j])
            set_cell_style_by_export(cell)

    # 结尾
    sheet.row_dimensions[3 + idx + 1].height = "25"
    sheet.cell(row=3 + idx + 1, column=1, value="送货人：")
    sheet.cell(row=3 + idx + 1, column=5, value="收货人：")

    # 保存
    wb.save(supply_file)
    wb.close()
    return supply_file


def set_print_fmt(supply: str) -> None:
    """
    设置打印格式
    :param supply: 打印文件路径
    """
    wb = openpyxl.load_workbook(supply)
    sheet = wb.active

    # 获取需要打印的范围
    print_area = sheet.dimensions
    # 设置打印范围
    sheet.print_area = print_area

    # 关闭
    wb.save(supply)
    wb.close()


def generate_supply_order(supply_name, supply_data, supply_file, to_dir):
    sort_data = sorted(supply_data, key=lambda x: x.get("外部订单号"))
    no_list = set([data.get("外部订单号") for data in sort_data])
    supply_list = []
    for no in no_list:
        # 生成文件名
        out_file_name = "{}供货单{}.xlsx".format(supply_name, no)
        out_path = os.path.join(to_dir, out_file_name)
        shutil.copyfile(supply_file, out_path)
        # 生成供应单并把文件地址加入供应单列表
        supply_list.append(generate_order_by_supply(no, supply_data, out_path, to_dir))
    # NEW~ 设置打印区间？
    [set_print_fmt(supply) for supply in supply_list]
    return supply_list


def set_cell_style_by_export(cell):
    # 字体
    font = Font(size=11, vertAlign='baseline')
    # 边框
    side = Side(border_style='thin', color='000000')
    border = Border(left=side, right=side, top=side, bottom=side)
    # 对齐
    align = Alignment(horizontal='center', vertical='center')
    # cell
    cell.font = font
    cell.border = border
    cell.alignment = align
    return cell


def make_storage_template(tool, template, supply_file, to_dir):
    # 1、拿去BOM数据
    df = pd.read_excel(tool, sheet_name="BOM")
    # 2、制作数据透视表
    df_pivot = pd.pivot_table(df, index=["子件料号二级", "半成品库存", "供应商", "库位"],
                              values=["分料数量"],
                              aggfunc={"分料数量": np.sum},
                              margins=False)
    # 3、按供应商分类数据
    export_data = []
    # 供货单文件字典
    supply_list = dict()
    for k, v in abbr_dict.items():
        supplier_type = k
        # 供货商供货单名称
        zip_name = "供应商{}供货单".format(supplier_type)
        # NEW~ 王子新材你是特殊的
        if k == "王子新材":
            supplier_type = "王子"
        group_data = df_pivot.query("供应商 == ['{}']".format(supplier_type))
        # 增加一个参照数据
        refer_data = pd.read_excel(tool, sheet_name="共用料号", usecols="A,E")
        tmp = subtotal_data(k, v, group_data, refer_data)
        # 生成供应单
        file_list = generate_supply_order(k, tmp, supply_file, to_dir)
        # NEW~ 将单一供货商的供货单据打包发送 PS: foxmail罪大恶极！！！
        to_zip = os.path.join(to_dir, "zip")
        if not os.path.exists(to_zip):
            os.makedirs(to_zip)
        zip_f = files_zip(to_zip, zip_name, file_list)
        supply_list.setdefault(k, [zip_f])
        export_data.extend(tmp)
    sort_data = sorted(export_data, key=lambda x: x.get("外部订单号"))
    # 导入模板
    load_template(template, sort_data)
    return supply_list


def get_formula_list(row):
    # 整理公式
    formulas = []
    # 合计公式
    total = "=SUM(B{num}:O{num})".format(num=row)
    formulas.append(total)
    # 成品库存
    finish_inventory = "=IFERROR(VLOOKUP(A{num},原材料、成品库存透视表!$A$1:$C$1000,2,0),0)".format(num=row)
    formulas.append(finish_inventory)
    # 1-4订单 -修正为-> 分料数量，不再分开发送
    order_formula = "=MAX(IF(ISERROR(VLOOKUP(A{num},固定需求料号!$A$1:$B$16,2,0)*14),(P{num})," \
                    "VLOOKUP(A{num},固定需求料号!$A$1:$B$16,2,0)*14)-Q{num})".format(num=row)
    formulas.append(order_formula)
    return formulas


def get_excel_len(in_sheet, is_col):
    if is_col:
        return len([i for i in in_sheet[1] if i.value is not None])
    else:
        return len([i for i in in_sheet if i[0].value is not None])


def compare_data_process(to_file, from_file):
    # 1、对被连个数据源的行数
    to_wb = openpyxl.load_workbook(to_file)
    from_wb = openpyxl.load_workbook(from_file)
    to_sheet = to_wb["需求"]
    from_sheet = from_wb.active
    # 获取各自最大行数
    to_len = to_sheet.max_row
    from_len = get_excel_len(from_sheet, False)
    if to_len > from_len:
        # 2、多余的行删除
        print("删除叫料表中多余的行数")
        to_sheet.delete_rows(from_len + 1, to_len)
    elif to_len < from_len:
        # 3、缺少的行补充
        print("补充叫料表中缺少的行数")
        diff_row = from_len - to_len
        sign_col_idx = 15
        for i in range(0, diff_row):
            current_idx = to_len + i + 1
            formulas = get_formula_list(current_idx)
            for j in range(0, len(formulas)):
                cell = to_sheet.cell(row=current_idx, column=sign_col_idx + j + 1, value=formulas[j])
                set_cell_style_by_export(cell)
    else:
        print("叫料表和附件数据表行数相等")
    # 关掉链接
    to_wb.save(to_file)
    to_wb.close()
    from_wb.close()


def data_close_by_sheet(target_file: str, sheet_name: str) -> None:
    # 获取表
    wb = openpyxl.load_workbook(target_file)
    sheet = wb[sheet_name]
    # 删除数据
    sheet.delete_cols(1, 2)
    wb.save(target_file)
    wb.close()


def statistical_data(files, in_dirs):
    """
    $统计中心：主方法
    :param files: 使用的文件映射表
    :param in_dirs: 使用的文件加映射
    :return: supply_file_dict 供应商单地址映射
    """
    # 校验文件稳定性
    (is_exist, msg) = check_file_exist(*list(files.values()))

    if is_exist:
        # 1、首先获取叫料工具表格
        material_tool = files.get("material")
        # 2、获取附件表格进行处理
        attach_excel = files.get("attach")
        # NEW~ 2.1 新增将附件表格与叫料表格进行对比
        compare_data_process(material_tool, attach_excel)
        # 拷贝表格的数据
        copy_excel_data(attach_excel, material_tool, "需求", copy_area="B,D:Q")
        # 3、获取系统库存导出数据进行处理
        export_excel = files.get("export")
        # 进行库存数据清空
        data_close_by_sheet(material_tool, "系统库存导出数据")
        # 同样也是拷贝表格的数据
        copy_excel_data(export_excel, material_tool, "系统库存导出数据", copy_area="B,T")
        # 获取已叫料但未收货数据
        not_receipt_excel = files.get("not_receipt")
        copy_excel_data(not_receipt_excel, material_tool, "系统库存导出数据", from_row=1, is_cover=False, copy_area="D,E")
        # 4、对工具表格进行刷数据
        excel_refresh(material_tool)
        excel_refresh(material_tool)
        # 5、制作导入模板和供货单
        template_file = files.get("template")
        supply_file = files.get("supply")
        to_dir = in_dirs.get("supply_order")
        # 开始生成模板和文件
        supply_file_dict = make_storage_template(material_tool, template_file, supply_file, to_dir)
        # 将供应商和映射文件列表返回
        return supply_file_dict
    else:
        raise FileNotFoundError("程序异常-缺少必要文件无法执行|<{}>".format(msg))


print(statistical_data(file_dict, dirs))

# ===============需求数据写入=============== #
# pf = pd.read_excel(r"C:\Users\Administrator\Downloads\测试包\20221209_SAP_97506873.xlsx", sheet_name= usecols="B,D:Q")
# book = load_workbook(r"C:\Users\Administrator\Downloads\原材料需求叫料11.30.xlsx")
# sheet = book["需求"]
# # 处理一下数据将列数据和body数据进行整合
# copy_data = [np.array(pf.columns)]
# copy_data.extend(pf.values)
# idx = len(copy_data)
# for i in range(0, idx):
#     for j in range(0, len(copy_data[i])):
#         sheet.cell(row=i+1, column=j+1, value=copy_data[i][j])
# book.save(r"C:\Users\Administrator\Downloads\原材料需求叫料11.30.xlsx")
# ======================================== #
