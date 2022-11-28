#!/usr/bin/env python
# -*- coding: UTF-8 -*-

# Description 交易明细到入模板
from util.config_util import get_config
import os
import decimal
from datetime import datetime
import pandas as pd
import shutil
import openpyxl

# ARGS

config_dict = get_config().get("config")
# 读取中出现异常被过滤出来的数据
ERR_DATA = []


# START:

# Tips: 获取当前运行文件夹
def get_current_runtime_dir():
    """
    获取当前运行文件夹
    :return: current_run_dir 当前运行文件夹
    """
    # 1 获取流程输出目录
    output_path = os.path.join(config_dict.get('home_dir'), 'output')
    # 2 获取当前运行的输出目录
    current_run_date = datetime.now().strftime('%Y%m%d')
    current_run_dir = os.path.join(output_path, current_run_date)
    # 3 返回当前运行文件夹
    return current_run_dir


# Tips: 获取交易详情文件列表
def get_transaction_detail():
    """
    获取到交易详情文件列表
    :return: details_file_list 详情文件地址列表
    """
    # 1 获取当前运行文件夹
    current_run_dir = get_current_runtime_dir()
    # 2 获取输出目录中的交易明细文件列表
    transaction_details_dir = os.path.join(current_run_dir, 'transaction_details')
    details_file_list = os.listdir(transaction_details_dir)
    # 3、将数据整理并返回
    details_file_list = \
        list(map(lambda details_file: os.path.join(transaction_details_dir, details_file), details_file_list))
    return details_file_list


# Tips: 根据文件路径进行处理交易详情数据
def process_transaction_detail(detail_file_path):
    """
    根据文件路径进行处理交易详情数据
    :param detail_file_path: 交易详情文件路径
    :return: details_data 经过初步筛选的数据
    """
    # 检测文件路径是否正确
    if not os.path.exists(detail_file_path):
        raise FileNotFoundError("程序异常-交易详情文件异常: 文件路径不存在 %s" % detail_file_path)
    # 设置需要的数据列表
    REQUIRED_COLUMN = [
        "交易结束时间",
        "交易流水号",
        "交易金额",
        "商户数据包",
        "商品名称",
        "支付方式"
    ]
    # 读取详情表格文件
    pf = pd.read_excel(detail_file_path)
    # 过滤出交易状态为支付成功的实际
    detail_data = pf.loc[pf["交易状态"] == "支付成功", REQUIRED_COLUMN]
    # 将数据进行返回
    return detail_data


# Tips: 客户数据包处理
def node_process_merchant_data_package(data_package):
    """
    客户数据包分割处理
    :param data_package: 数据包数据
    :return: (company_name: 单位名称, business_code: 业务编码)
    """
    # 获取数据包分割标识
    flags = config_dict.get("data_package_delimiter").split("|")
    package_err_flag = False
    flag_by_type = ""
    for flag in flags:
        if data_package.count(flag) == 1:
            package_err_flag = True
            flag_by_type = flag
    if not package_err_flag:
        raise Exception("程序异常-商户数据包解析异常：没有标识符分割 <{}>".format(data_package))
    data_list = data_package.split(flag_by_type)
    company_name = data_list[0]
    business_code = data_list[1]
    return company_name, business_code


# Tips: 处理单行数据 => entity
def node_process_by_row(row, pay_type):
    """
    处理单行数据为字典
    :param row: 行数据
    :param pay_type: 支付类型
    :return: row_entity 导入数据行
    """
    # 取出单行数据
    trade_end_time = getattr(row, "交易结束时间")
    trade_serial_no = getattr(row, "交易流水号")
    trade_amount = getattr(row, "交易金额")
    product_name = getattr(row, "商品名称")
    merchant_data_package = getattr(row, "商户数据包")
    # 处理数据细节
    trade_end_time = datetime.strptime(trade_end_time, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
    # 获得单位名称和业务编码
    try:
        (company_name, business_code) = node_process_merchant_data_package(merchant_data_package)
    except Exception as e:
        # 记录异常数据行的【交易流水号】
        ERR_DATA.append(trade_serial_no)
        raise Exception("数据异常：{}".format(e))
    # 组装摘要
    tmp_date = trade_end_time.split("-")
    _cn = ["年", "月", "日"]
    fmt_data = "".join(["".join(list(tmp)) for tmp in zip(tmp_date, _cn)])
    summary = "{}{}{}{}".format(fmt_data, company_name, product_name, business_code)
    # 组装实体
    row_entity = {
        "*公司_编码": config_dict.get("company_code"),
        "*币别_编码": config_dict.get("currency_code"),
        "*银行账户_银行账号": config_dict.get("bank_account"),
        "*交易日期": trade_end_time,
        "*交易时间": trade_end_time,
        "业务类型": "",
        "票据号_l2": "",
        "*业务参考号": trade_serial_no,
        "付款金额": "",
        "收款金额": trade_amount,
        "余额": "0",
        "摘要": summary,
        "对方单位": company_name,
        "对方银行账户": "",
        "对方银行": "",
        "银行上划": "",
        "银行下拨": "",
        "流水号": "",
        "对账标识码": "",
        "业务分类": pay_type,
    }
    return row_entity


# Tips: 将表格数据转换为简易列表
def node_process_by_pay(pay_type, pay_data):
    """
    将分类数据从dataframe转换为字典数据
    :param pay_type: 支付类型
    :param pay_data: 该支付类型数据
    :return: 该支付类型数据字典 => {pay_type: [{}...{}]}
    """
    data_list = []
    for row in pay_data.itertuples():
        try:
            data_list.append(node_process_by_row(row, pay_type))
        except Exception as e:
            # 跳过该数据
            print(e)
            continue
    return {pay_type: data_list}


# Tips: 对经过初步筛选的详情数据集合进行深度加工
def detail_deep_process(detail_data_list):
    """
    将过滤【支付成功】后的数据加工为模板文件格式数据
    :param detail_data_list: 交易详情过滤数据
    :return: input_datas 模板文件格式数据集合
    """
    # 0 将数据集合合并成为一个DataFrame对象
    detail_data = pd.concat(detail_data_list, ignore_index=True)
    # 1 首先以支付方式过滤出支付方式
    group_data = detail_data.groupby("支付方式")
    # 获取支付方式类别列表
    payment_types = list(group_data.groups.keys())
    # 组合原始数据字典
    group_dict = \
        dict([(payment_type, group_data.get_group(payment_type).reset_index()) for payment_type in payment_types])
    print(group_dict)
    # 2 分别解析每种支付方式的各数据
    input_datas = []
    for key, value in group_dict.items():
        # 将表格数据转换为简易列表并存入数组
        input_datas.append(node_process_by_pay(key, value))
    # 返回数据集合
    return input_datas


# Tips: 获取导入模板
def get_import_template():
    """
    获取导入模板文件地址
    :return: template_file_path 模板文件地址
    """
    # 1 获取流程配置目录
    input_path = os.path.join(config_dict.get('home_dir'), 'input')
    # 2 获取导入模板文件夹
    template_file_path = os.path.join(input_path, '银行账户交易明细导入模板.xlsx')
    # 3 返回模板文件路径
    return template_file_path


# Tips: 处理交易数据
def process_transaction_data():
    """
    将交易数据处理为可以被导入文件的数据集合
    :return: process_data 处理后的数据集合
    """
    # 获取当前明细文件列表
    detail_path_list = get_transaction_detail()
    if not detail_path_list:
        raise FileNotFoundError("程序异常-配置文件异常: 交易详情文件下载异常")
    # 处理方法
    # 将文件列表中的文件进行初步筛选存入列表中
    detail_list = [process_transaction_detail(path) for path in detail_path_list]
    # 对数据集合进行深加工 => 获得一个包含本次执行的所有数据的dict集合 [{},{},...{}]
    process_data = detail_deep_process(detail_list)
    return process_data


# Tips: 在数据后增加类型汇总
def sum_data_by_type(pay_type, data):
    """
    按照支付类型聚合总金额数据并加入数据集合
    :param pay_type: 支付类型
    :param data: 该类型数据集合
    :return: data 增加了聚合数据的数据集合
    """
    # 总金额
    sum_amount = decimal.Decimal('0.0')
    # 计算该类型总金额
    for entity in data:
        sum_amount = sum_amount + decimal.Decimal(str(entity.get("收款金额")))
    # 深拷贝先前集合中的一条数据
    sun_data = data[-1].copy()
    # 支付方式缩写映射
    abbr_map = {
        "微信": 'WX',
        "支付": 'ZFB',
        "银联": 'YL'
    }
    # 将聚合数据填入先前拷贝出来的数据中
    context = "{}转出".format(pay_type)
    end_time = datetime.strptime(sun_data.get("*交易日期"), "%Y-%m-%d").strftime("%Y%m%d")
    trade_serial_no = "{}{}{}".format(config_dict.get("company_code"), abbr_map.get(pay_type[:2]), end_time)
    sun_data["*业务参考号"] = trade_serial_no
    sun_data["付款金额"] = float(sum_amount)
    sun_data["收款金额"] = ""
    sun_data["摘要"] = context
    sun_data["对方单位"] = context
    sun_data["业务分类"] = ""
    # 将数据加入原先数据集
    data.append(sun_data)
    # 返回处理好的数据集合
    return data


# Tips: 加载数据
def load_data(file_path, data):
    """
    将数据保存在模板文件中
    :param file_path: 文件地址
    :param data: 加载数据
    """
    # 读取模板列数据
    pf = pd.read_excel(file_path, header=3)
    # 将数据包分类型注入dataframe
    for sort in data:
        for pay_type, details in sort.items():
            print("当前加载的是{}类型".format(pay_type))
            # 增加汇总数据
            type_data = sum_data_by_type(pay_type, details)
            # 将数据插入空的模板模型中
            for ioc in type_data:
                pf.loc[len(pf)] = ioc
    print("最终形态：\n")
    print(pf)
    # 使用openpyxl写入模板文件
    wb = openpyxl.load_workbook(file_path)
    sheet = wb['交易明细']
    # 获取数据数组 Ps: 这里有点混乱
    out_values = pf.values
    idx = len(out_values)
    # 写入
    for i in range(0, idx):
        for j in range(0, len(out_values[i])):
            sheet.cell(row=4 + i + 1, column=j + 1, value=out_values[i][j])
    # 保存
    wb.save(file_path)


# Tips: 将模板文件拷贝到当前运行目录中
def copy_template_file():
    """
    将模板文件拷贝到当前按运行目录中
    :return: to_file_path 拷贝文件当前路径
    """
    template_file_path = get_import_template()
    current_run_date = datetime.now().strftime('%Y-%m-%d')
    to_file_name = "银行账户交易明细导入{}.xlsx".format(current_run_date)
    to_file_path = os.path.join(get_current_runtime_dir(), to_file_name)
    shutil.copyfile(template_file_path, to_file_path)
    return to_file_path


# Tips: 数据导入模板
def data_import_template():
    """
    获取交易明细数据填入导入模板文件
    :return: to_file_path 返回填入的导入模板文件地址
    """
    # 0 将模板文件拷贝到当前运行目录中
    to_file_path = copy_template_file()
    # 1 获取将要导入的数据
    input_data = process_transaction_data()
    # 2 将数据导入到模板中
    load_data(to_file_path, input_data)
    # 3 返回导入文件地址
    return to_file_path


# END;

# TEST?>

print("导入模板生成路径：<{}>".format(data_import_template()))
# print(process_transaction_data())
# REQUIRED_COLUMN = [
#     "交易结束时间",
#     "交易流水号",
#     "交易金额",
#     "商户数据包",
#     "商品名称",
#     "支付方式"
# ]
#
# test_path = r'C:\RPA\招商_招投标中心_交易明细导入导出\output\20221125\transaction_details\订单交易明细2022-11-22-10-57.xlsx'
# table = pd.read_excel(test_path)
# pro_table = table.loc[table["交易状态"] == "支付成功", REQUIRED_COLUMN]
# group = pro_table.groupby(["支付方式"]).get_group("支付宝").reset_index()
# #
# for row in group.itertuples():
#     print(type(getattr(row, "交易金额")))

# d = pd.read_excel(r'E:\银行账户交易明细导入模板.xlsx', header=3)
#
# bb = {
#     "*公司_编码": "01590002",
#     "*币别_编码": "RMB",
#     "*银行账户_银行账号": "XN-30899917278000B",
#     "*交易日期": "2022-11-15",
#     "*交易时间": "2022-11-15",
#     "业务类型": "",
#     "票据号_l2": "",
#     "*业务参考号": "100422111508292660511855",
#     "付款金额": "",
#     "收款金额": "500.22",
#     "余额": "0",
#     "摘要": "2022年11月15日中交二航局建筑科技有限公司平台服务费GX2022111508291223632877827",
#     "对方单位": "中交二航局建筑科技有限公司",
#     "对方银行账户": "",
#     "对方银行": "",
#     "银行上划": "",
#     "银行下拨": "",
#     "流水号": "",
#     "对账标识码": "",
#     "业务分类": "银联二维码",
# }
#
# wb = openpyxl.load_workbook(r'E:\银行账户交易明细导入模板.xlsx')
# sheet = wb['交易明细']
# print(sheet.max_row)
# d.loc[len(d)] = bb
# d.loc[len(d)] = bb
# d.loc[len(d)] = bb
# value1 = d.values
# index = len(value1)
# for i in range(0, index):
#     for j in range(0, len(value1[i])):
#         sheet.cell(row=4 + i + 1, column=j + 1, value=value1[i][j])
#
# wb.save(r'E:\银行账户交易明细导入模板.xlsx')

#
# rows = dataframe_to_rows(d)
# for row in rows:
#     for col in row:
#         print(col)
# print(d)
# with pd.ExcelWriter(r'E:\银行账户交易明细导入模板.xlsx', mode='a') as writer:
#     d.to_excel(writer, sheet_name="sheet1", header=False, index=False, startrow=4)
