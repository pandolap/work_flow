# -*- coding: utf-8 -*-
import time

import openpyxl
import json
import os

excel_area = "A1:J3"
data_list_str = r'D:\RPA\缴税\数据源\6.xlsx'
invoice_dir = r"D:\RPA"
log_path = r"D:\RPA\logs"

# 打开清单表格文件
wb = openpyxl.load_workbook(data_list_str, data_only=True)
# 读表格文件
sheet_list = wb.sheetnames
# 获取首个Sheet页
sh = wb[(sheet_list[0])]
# 获取列宽以及行宽
row_max = sh.max_row
cel_max = sh.max_column

table = []
bill_list = []
# 获取截取范围
area = excel_area.split(":")
for row in sh[area[0]:area[1]]:
    table.append([cell.value for cell in row])
is_start = False
# 获取值
for index, r in enumerate(table):
    try:
        if '车架号' in r[0]:
            is_start = True
            continue
    except TypeError:
        pass
    if is_start:
        bill = {}
        bill.setdefault("车架号", r[0])
        bill.setdefault("报关单号", r[5])
        bill.setdefault("合同号", r[2])
        bill.setdefault("缴税日期", r[9])
        bill.setdefault("贵司合同号", r[5])
        bill.setdefault("总税款", r[6])
        bill.setdefault("单据信息", "")
        bill_list.append(bill)
