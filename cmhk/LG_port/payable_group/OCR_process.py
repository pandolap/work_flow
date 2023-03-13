#!/usr/bin/env python
# -*- coding: UTF-8 -*-

import json
import os
import re
import traceback
import hmac
import time
import base64
import hashlib

DEBUG = __name__ == '__main__'
ocr_list = []

if DEBUG:
    dir_name = r'C:\Users\Administrator\Downloads\15080005-GXAP-20230213-0002'
    flow_json = os.path.join(dir_name, '审批流.json')
    ocr_json = os.path.join(dir_name, 'ocr.json')
    with open(flow_json, 'r', encoding='utf-8') as f:
        data = json.load(f)

    data['data']['advice'] = ''
    data['data']['verifyResult'] = ''

    with open(ocr_json, 'r', encoding='utf-8') as f:
        ocr_sub_list = json.load(f)
        for item in ocr_sub_list:
            # ['response']['data']['identify_results']
            d = {
                'response': {
                    'data': {
                        'identify_results': item
                    }
                }
            }
            ocr_list.append(d)
else:
    with open(
            r"C:\Users\Administrator\Downloads\4050-GXAP-20230210-0001\4050-GXAP-20230210-0001_pre.json") as f:
        flow = f.read()
    data = json.loads(flow)

log_dir = data['config'].get('log_dir')
home_dir = data['config'].get('home_dir')
No = data['data']['baseInfo']['单据编号']
invoice_dir = os.path.join(os.path.join(log_dir, 'ocr与审批流运行数据'), No)
if not os.path.exists(invoice_dir):
    os.makedirs(invoice_dir)
baseInfo = data['data']['baseInfo']

pre_log_file = os.path.join(invoice_dir, No + '_pre.json')
with open(pre_log_file, 'w', encoding='utf-8') as f:
    json.dump(data, f)

# Todo url_dict_new 是所有图片的链接 return result.images.N
url_dict_new = data['data'].get('imgN', [])

# 记录 ocr时间
from datetime import datetime

now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
if not os.path.exists(log_dir):
    os.makedirs(log_dir)
with open(os.path.join(log_dir, 'OCR记录.txt'), 'a', encoding='gbk') as f:
    record = ['[', now, ']', '图片数量%s' % len(url_dict_new), '\n']
    f.write(''.join(record))

# Todo advice 是之前所有的意见
advice = data['data'].get('verifyResult', '')
flow_advice = advice
inside_pattern = re.compile(r'([^【】]*)(【提示：机器人试运行，(【[^【】]*】[^【】]*)】)(.*)')
inside_pattern_rs = inside_pattern.search(advice)
if inside_pattern_rs:
    advice = inside_pattern.sub(r'\1\3\4', advice)

# Todo costBearingcompany 是费用承担公司 深圳妈港仓码有限公司
costBearingcompany = baseInfo.get('公司')

# Todo receivingBank 是收款银行 中国工商银行股份有限公司深圳梅林支行
receivingBank = baseInfo.get('收款银行')

# Todo payee 是收款人  深圳柏汇自动化设备有限公司
payee = baseInfo.get('收款人')

# Todo collectionAccount 收款账户   4000026209200366760
collectionAccount = baseInfo.get('收款账户')

# Todo moneyFee 合计税额   106.99
moneyFee = baseInfo.get('合计税额')

# Todo feeSum 应付金额  	0
feeSum = baseInfo.get('应付金额')

# Todo finalPay 付款金额
finalPay = baseInfo.get('付款金额', '')

import requests
import re
from requests_toolbelt import MultipartEncoder
import openpyxl
import pandas as pd
import numpy as np
import uuid


def check_field(standard: str, tested: str) -> tuple:
    if not (standard and tested):
        raise Exception("比较字符串时，入参为空")
    # 对两个字段进行去空格处理
    standard = standard.strip()
    tested = tested.strip()
    # 将两个字段进行逐字比对
    # 首先比较字段的长度
    standard_len = len(standard)
    tested_len = len(tested)
    idx = 1
    is_flag = False
    if standard_len != tested_len:
        is_flag = True
    compare_len = min(standard_len, tested_len)
    for i in range(0, compare_len):
        if standard[i] != tested[i]:
            is_flag = True
            break
        else:
            idx += 1
    return is_flag, idx


def imageJson(url, count=5, ocr_result=None, img_content=None):
    if count < 0:
        invoice_err_pic = os.path.join(invoice_dir, 'invoice_%s.jpg' % uuid.uuid1())
        invoice_err_record = os.path.join(invoice_dir, '出错图片记录.txt')
        if img_content and ocr_result:
            with open(invoice_err_pic, 'wb') as fb:
                fb.write(img_content)
        info_list = [invoice_err_pic, json.dumps(ocr_result)]

        with open(invoice_err_record, 'a', encoding='utf-8') as f:
            f.write(','.join(info_list) + '''\n''')
        return ocr_result
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/"
                      "537.36 (KHTML, like Gecko) Chrome/75.0.3770.80 Saf"
                      "ari/537.36",
    }
    image_content = requests.get(url, headers=headers, timeout=15).content

    m = MultipartEncoder(
        fields={
            'file': ('images.jpg', image_content, 'text/plain'),
            'uuid': '2563f6b39b6d42c29d28ec3eb8e1d00f',
            'checkInvoice': 'True',
            'ifOcrType': 'False',
        }
    )
    try:
        omg = requests.post(url="http://ocr.pre.ninetechone.com/fileOCRAndcheckInvoice", data=m,
                            headers={'Content-Type': m.content_type})
        # return omg.json()
        rs = omg.json()
        if rs.get('message', '').find('error') > -1:
            return imageJson(url, count=count - 1, ocr_result=rs, img_content=image_content)
        if rs.get('error', 0) > 0:
            if rs.get('message', '') == 'Invalid imageFile parameter, maxsize 8MB':
                return rs
            return imageJson(url, count=count - 1, ocr_result=rs, img_content=image_content)
        return rs
    except:
        count = count - 1
        info = traceback.format_exc()
        return imageJson(url, count, ocr_result=info)


# Todo url_dict_new 是所有图片的链接 return result.images.N
# Todo advice 是之前所有的意见
# Todo costBearingcompany 是费用承担公司 深圳妈港仓码有限公司
# Todo receivingBank 是收款银行 中国工商银行股份有限公司深圳梅林支行
# Todo payee 是收款人  深圳柏汇自动化设备有限公司
# Todo collectionAccount 收款账户   4000026209200366760
# Todo moneyFee 合计税额   106.99
# Todo feeSum 应付金额  	0
totalSum = 0
taxSum = 0
listExcel = []
advice_first = advice

bearing_flag = None  # 费用承担公司标识
taxnumber_flag = None  # 税号标识
seal_flag = None  # 盖章标识
formname_flag = None  # 发票联标识
receivebank_flag = None  # 收款银行标识
seller_flag = None  # 销售方标识
collectionAccount_flag = None  # 收款账户标识
flight_flag = None  # 机票仓位标识
seat_flag = None  # 货车仓位标识
question_list = []
log_list = []
seller_bank_account_flag = None
# 抵扣联判断列表
credit_union_flag = False
# 开票时间列表
invoice_time_error_flag = False
# 付款金额(总)
pay_sum = 0
# 合计税额(总)
tax_sum = 0
# 是否有发票
invoice_exists = False

# [暂缓]新增是否有中选或是备案
# selected_filed = False
# ====对法务信息进行处理====

SIGN_SK = '5e278fb467e94d2e9c9241183c6ac0c5'
SIGN_AK = '27e28e8ea93b44d69c21d39189a7e775'
# SIGN_AK = 'ddec77bb168e465c96a970227889e243'
# SIGN_SK = '37ff70b260764c56ad489e847387b262'

DEBUG = __name__ == "__main__"


def record_runtime_data(runtime_data, name):
    with open(os.path.join(invoice_dir, name) + '.json', 'w', encoding='utf-8') as f:
        json.dump(runtime_data, f)


def get_header() -> dict:
    timestamp = int(time.time() * 1000)
    data = '\ntimestamp:' + str(timestamp)
    hashing = hmac.new(bytes(SIGN_SK, encoding='utf-8'), bytes(data.strip('&'), encoding='utf-8'),
                       hashlib.sha1).digest()
    sign = base64.b64encode(hashing)
    headers = {'timestamp': str(timestamp), 'Authorization': SIGN_AK + ':' + bytes.decode(sign)}
    return headers


def get_images_result(asset_code: str) -> dict:
    """ 获取影像图片信息 """
    url1 = f"https://gip.cmft.com/gip-api/v1/images/{asset_code}"
    # url1 = f"https://gip-st1.uat.cmft.com:8085/gip-api/v1/images/{asset_code}"
    headers = get_header()
    res = requests.get(url1, headers=headers)
    if res.status_code == 200:
        return res.json()
    return {"识别失败": res.text}


def get_ocr_text(img_url: str) -> dict:
    if DEBUG:
        request_url = "https://openapi.cmft.com/gateway/general1/1.0.0/api/pdfrec"
    else:
        request_url = "http://openapi.cmft.com:8080/gateway/general1/1.0.0/api/pdfrec"

    payload = MultipartEncoder(
        fields={
            "file": ("img.jpg", requests.get(img_url).content),
            "content": "multipart/form-data"
        }
    )

    print(payload.content_type)
    header = {
        "Authorization": "ACCESSCODE 142717D9360287ED5BD1D03E1B6805D2",
        "x-Gateway-APIKey": "b9180959-bb39-4f58-935a-bacc6a80d16c",
        "Content-Type": payload.content_type,
    }

    response = requests.post(request_url, data=payload, headers=header)
    return response.json()


def get_original_num(title_text: str) -> str:
    """
    获取原单编号
    :param title_text: 当前单据标题
    :return: 原单编号
    """
    begin_idx = title_text.find("原单编号")
    before_idx = title_text.rfind(")")
    if begin_idx != -1 and before_idx != -1:
        _num = title_text[begin_idx:before_idx]
        res_ = re.findall(r"\d+-\w+-\d+-\d+", _num)
        if len(res_) > 0:
            original_num = res_[0]
            return original_num
        else:
            return ""
    else:
        return ""


def get_record_info(rx_no):
    if rx_no:
        res_out = get_images_result('E{}'.format(rx_no))
        url_list = []
        if res_out.get("code") == "Y":
            url = None
            res_body = res_out.get("body")
            images_data = res_body.get("images")
            for image in images_data:
                for img in image.get("sameTypeList"):
                    if img.get("treeId", "A004") == "A006":
                        url = img.get("url")
                        url_list.append(url)
            if len(url_list) > 0:
                return True, url_list
            else:
                return False, None
        else:
            return False, None
    else:
        return False, None


def forecast_match_amount(arg):
    res_amount = 0.0
    if arg:
        args = re.findall(r'\d+', arg)
        if len(args) == 1:
            res_amount = float(args[0])
        else:
            mantissa = len(args[-1])
            if mantissa == 2:
                # 正常小数位
                res_amount = float('{}.{}'.format(''.join(args[:-1]), args[-1]))
            else:
                # 拼接整数
                res_amount = float(''.join(args))
    return res_amount


def get_image_amount(url):
    res = get_ocr_text(url)
    if res.get('code', 'N') == 'Y':
        is_selected = None
        ocr_detail = res.get("details", "")[0]
        lines = ocr_detail.get("context").get("lines")
        for line in lines:
            if "中选" in line:
                is_selected = True
                break
            if "备案" in line:
                is_selected = False
                break
        # 成交金额
        deal_amount = 0
        # 预算金额
        budget_amount = 0
        if is_selected:
            record_runtime_data(res, '中选通知书OCR识别记录')
            # 中选
            is_budget = False
            is_deal = False
            for line in lines:
                if "成交金额" in line:
                    is_deal = True
                if "预算价" in line:
                    is_budget = True
                if is_deal:
                    if re.findall(r"\d+", line):
                        deal_amount = forecast_match_amount(line)
                        is_deal = False
                if is_budget:
                    if re.findall(r"\d+", line):
                        budget_amount = forecast_match_amount(line)
                        is_budget = False

        else:
            record_runtime_data(res, '备案通知书OCR识别记录')
            # 备案
            is_budget = False
            is_deal = False
            for line in lines:
                if '预算价' in line:
                    is_budget = True
                if "合同金额" in line:
                    is_deal = True
                if is_deal:
                    if re.findall(r"\d+", line):
                        deal_amount = forecast_match_amount(line)
                        is_deal = False
                if is_budget:
                    if re.findall(r"\d+", line):
                        budget_amount = forecast_match_amount(line)
                        is_budget = False

        return is_selected, deal_amount, budget_amount
    else:
        return None, None, None


def amount_update(amount):
    try:
        res = float(amount)
    except Exception:
        # 寻找金额
        is_decimal_point = False
        if ',' in amount:
            amount = amount.replace(',', '')
        if '，' in amount:
            amount = amount.replace('，', '')
        if '.' in amount:
            is_decimal_point = True
            amount = amount.replace('.', '')
        if is_decimal_point:
            amount = "{}.{}".format(amount[:-2], amount[-2:])
        res = float(amount)
    return res


def get_amount(org_str):
    return org_str.replace(',', '')


# 获取单据原单号
title_text = data['data']['current_title']
# 区分单据获得：
if '自动扣款' in title_text:
    original = No
else:
    original = get_original_num(title_text)
(selected_filed, ocr_list_1) = get_record_info(original)
deal = None
if selected_filed:
    (is_select, deal, budget) = get_image_amount(ocr_list_1[0])
    deal = deal
    budget = budget
    # 比较
    if is_select == True:
        # 中选
        pass
    elif is_select == False:
        # 备案
        if float(budget) == 0.0:
            advice += '【未提取到预算金额，请人工校验】；'
        else:
            if float(budget) >= float(deal):
                advice += '不超出预算；'
            else:
                advice += '【超出预算】；'
    else:
        # 山野没有
        selected_filed = False
else:
    advice += '【未找到中选或备案】；'
    # No
this_pay = baseInfo.get("本次支付金额", "")
total_pay = baseInfo.get("累计支付金额", "")

if this_pay == "" and total_pay == "":
    advice += '【未查到法务信息】；'
else:
    if this_pay == '':
        this_pay = '0'
    if total_pay == '':
        total_pay = '0'
    legal = float(get_amount(this_pay)) + float(get_amount(total_pay))
    advice += '含本单合计：{}元；'.format(str(round(legal, 2)))
    if selected_filed:
        if float(deal) < legal:
            advice += '【超额支付】；'
        else:
            advice += '未超额支付；'
    else:
        pass
        # advice += ''

# 2.13 物资采购付款申请单
annex_num = baseInfo.get('附件数')
if not annex_num:
    annex_num = 0

if int(annex_num) <= 5:
    check_list = url_dict_new[1:5]
    if len(check_list) == 4:
        is_word_exist = False
        for img_url in check_list:
            res_text = get_ocr_text(img_url)
            if res_text.get('code', 'N') == 'Y':
                ocr_detail = res_text.get("details", [])[0]
                title_text = ocr_detail.get("total_text", "")
                if title_text:
                    check_text = title_text[:50]
                    if '物资采购付款申请单' in check_text:
                        is_word_exist = True
                        break
        if is_word_exist:
            business_invoice_no = baseInfo.get('业务单据号')
            if business_invoice_no not in ['', ' ', '...']:
                advice += '业务单号完整；'
            else:
                advice += '【请核查业务单号】；'

# try:
# for i in url_dict['images']['N']:
for url_index, url in enumerate(url_dict_new):
    # url = i['url']  # 拿到每一条票据的url地址
    url_index += 1
    try:
        if not DEBUG:
            omg = imageJson(url)
        else:
            omg = ocr_list[url_index - 1]

        if omg.get('error') and omg.get('message', '') == 'Invalid imageFile parameter, maxsize 8MB':
            advice += "【第{}张影像图片大小超过8M，识别不了】".format(url_index)
            continue
        log_list.append(omg['response']['data']['identify_results'])
        for item in omg['response']['data']['identify_results']:
            type = item['type']
            if type == "0000":
                # advice = advice + "没有检测到发票；"
                pass
            if type == "1001" or type == "1002" or type == "1003" or type == "1004":  # 增值税普通发票
                invoice_exists = True
                # 2022/01/15 检查是否有抵扣联, 开票时间跟当前时间比较年份 --------------------------------------------------------
                invoice_item = item.get('details', dict())
                if invoice_item.get('form_name') == '抵扣联':
                    credit_union_flag = True
                try:
                    invoice_time = datetime.strptime(invoice_item.get('date'), '%Y年%m月%d日')
                except:
                    invoice_time = datetime.now()
                    advice += '【第%s张发票时间有遮挡】' % url_index
                now = datetime.now()
                if now.year > invoice_time.year:
                    invoice_time_error_flag = True

                # 总金额和税额比较
                # ocr:
                pay = invoice_item.get('total', '').replace(',', '')
                if pay == '':
                    advice += '【第%s张发票付款金额识别不出来】' % url_index
                    pay = 0
                else:
                    pay = round(float(pay), 2)

                pay_tax = invoice_item.get('tax', '').replace(',', '')
                if pay_tax == '':
                    advice += '【第%s张发票税额识别不出来】' % url_index
                    pay_tax = 0
                else:
                    pay_tax = round(float(pay_tax), 2)

                pay_sum += pay
                tax_sum += pay_tax

                # 2022/01/15 检查是否有抵扣联, 开票时间跟当前时间比较年份 --------------------------------------------------------

                # 购买方名称
                buyer = item['details'].get('buyer', '')
                if buyer == '':
                    advice += '第%s张发票购买方名称识别不出来' % url_index
                else:
                    buyer = buyer.replace('（', '').replace('）', '').replace('(', '').replace(')', '')

                # hlt添加、总分公司对应
                sub_company_col = 5
                father_company_col = 3
                company_dict = {}
                wb = openpyxl.load_workbook(os.path.join(home_dir, '辽港母分公司对应表.xlsx'))
                table = wb.get_sheet_by_name('Sheet1')
                max_row = table.max_row
                for irow in range(max_row):
                    irow += 1
                    relation = {table.cell(irow, sub_company_col).value: table.cell(irow, father_company_col).value}
                    company_dict.update(relation)
                if costBearingcompany in company_dict.keys():
                    costBearingcompany = company_dict[costBearingcompany]
                costBearingcompany = costBearingcompany.replace('（', '').replace('）', '').replace('(', '').replace(
                    ')', '')
                # advice = advice.replace("该单没有发票；","")
                buyer = buyer.replace(' ', '')
                costBearingcompany = costBearingcompany.replace(' ', '')
                if buyer == costBearingcompany:
                    print("购买方名称与费用承担公司一致；")
                    bearing_flag = 0
                else:
                    # advice = advice + "【购买方名称与费用承担公司不一致】；"
                    # bearing_flag = 1
                    (is_, diff_place) = check_field(buyer, costBearingcompany)
                    advice += '【请核查购买方名称，第{}个不同】；'.format(diff_place)
                    question_list.append(url_index)
                # 纳税人识别号
                buyer_tax_id = item['details'].get('buyer_tax_id', '')
                if buyer_tax_id == '':
                    advice += '【第%s张发票纳税人识别号识别不出来】' % url_index
                excel_data = pd.read_excel(os.path.join(home_dir, '辽港纳税人识别号.xlsx'))
                excelSerise = pd.Series(excel_data['企业名称'])
                excelNp = np.array(excelSerise)
                for value in excelNp:
                    listExcel.append(value)
                if buyer in listExcel:
                    if buyer_tax_id == excel_data['税号'][listExcel.index(buyer)]:
                        # advice = advice + "税号一致；"
                        taxnumber_flag = 0
                    else:
                        # advice = advice + "【税号不一致】；"
                        # taxnumber_flag = 1
                        advice += '【请核查税号，OCR显示{}】'.format(str(buyer_tax_id))
                        question_list.append(url_index)

                try:
                    # 发票盖章
                    company_seal = item['details']['company_seal']
                    if company_seal == "1":
                        # advice = advice + "第{}张影像中发票有盖章；".format(url_index)
                        seal_flag = 0
                    elif company_seal == "0":
                        # advice = advice + "【第{}张影像中发票没有盖章】；".format(url_index)
                        seal_flag = 1
                        question_list.append(url_index)
                except:
                    # advice = advice + "【第{}张影像中发票没有盖章】；".format(url_index)
                    seal_flag = 1
                    question_list.append(url_index)
                try:
                    # 发票联
                    form_name = item['details'].get('form_name')
                    if form_name:
                        # advice = advice + "第{}张影像中发票有发票联；".format(url_index)
                        formname_flag = 0
                    else:
                        # advice = advice + "【第{}张影像中发票没有发票联】；".format(url_index)
                        if type != '1003':  # 电子发票-普票可能没有”发票联“三个字，这是正常的
                            formname_flag = 1
                            question_list.append(url_index)
                        else:
                            formname_flag = 0
                except:
                    # advice = advice + "【第{}张影像中发票没有发票联】；".format(url_index)
                    if type != '1003':  # 电子发票-普票可能没有”发票联“三个字，这是正常的
                        formname_flag = 1
                        question_list.append(url_index)

            if type == "1001" or type == "1002" or type == "1003" or type == "1004":
                invoice_exists = True
                # 价税合计
                total = item['details'].get('total', 0)
                if total == '':
                    total = 0
                totalSum = totalSum + round(float(total), 2)
                # 销售方名称
                seller = item['details']['seller'].strip()
                try:
                    # 销售方开户行及账号
                    seller_bank_account = item['details']['seller_bank_account']
                    # 销售方开户行
                    try:
                        seller_bank = re.findall(r"\d+", seller_bank_account)[0].strip()
                    except:
                        seller_bank = seller_bank_account
                    if ":" in seller_bank:
                        seller_bank = seller_bank.replace(':', '')
                    # 银行关键字
                    bankCompany = ["中国银行", "中行", "农业银行", "农行", "工商银行", "工行", "建设银行", "建行", "招商银行", "招行"]

                    if seller_bank == receivingBank:
                        print("收款银行与发票影像中的销售方开户行一致")
                    else:
                        for bankItem in bankCompany:
                            if bankItem in seller_bank:
                                if bankItem in receivingBank:
                                    receivebank_flag = 0
                                    print("收款银行与发票影像中的销售方开户行一致")
                                else:
                                    # advice = advice + "第{}张影像中收款银行与发票影像中的销售方开户行不一致；".format(url_index)
                                    # receivebank_flag = 1
                                    (is_, diff_place) = check_field(seller_bank, receivingBank)
                                    advice += '【请核查收款银行，第{}个不同】；'.format(diff_place)
                                    question_list.append(url_index)
                    # 付款人母分公司对应
                    sub_company_col = 5
                    father_company_col = 3
                    company_dict = {}
                    wb = openpyxl.load_workbook(os.path.join(home_dir, '辽港母分公司对应表.xlsx'))
                    table = wb.get_sheet_by_name('Sheet1')
                    max_row = table.max_row
                    for irow in range(max_row):
                        irow += 1
                        relation = {table.cell(irow, sub_company_col).value: table.cell(irow, father_company_col).value}
                        company_dict.update(relation)
                    if payee in company_dict.keys():
                        payee = company_dict[payee]

                    payee = payee.replace('（', '').replace('）', '').replace('(', '').replace(')', '')
                    seller = seller.replace('（', '').replace('）', '').replace('(', '').replace(')', '')

                    if '月度预提单' not in data['data']['current_title']:
                        # 去掉当前的收款人和销售方名称校验
                        if payee == seller:
                            seller_flag = 0
                            print("第{}张影像中收款人与发票中的销售方名称一致".format(url_index))
                        else:
                            # advice = advice + "第{}张影像中收款人与发票中的销售方名称不一致；".format(url_index)
                            # seller_flag = 1
                            (is_, diff_place) = check_field(payee, seller)
                            advice += '【请核查收款人，第{}个不同】；'.format(diff_place)
                            question_list.append(url_index)

                    # if receivingBank == seller_bank:
                    #     print("收款银行与发票影像中的销售方开户行一致")
                    # else:
                    #     advice = advice + "收款银行与发票影像中的销售方开户行不一致；"

                    # 销售方账号
                    try:
                        seller_account = re.findall(r"\d+\.?\d*", seller_bank_account)[0]
                        if seller_account.strip().find(collectionAccount.strip()) > -1:
                            collectionAccount_flag = 0
                            print("收款账户与发票影像中的销售方账号一致")
                        else:
                            # advice = advice + "第{}张影像中发票收款账户与发票影像中的销售方账号不一致；".format(url_index)
                            # collectionAccount_flag = 1
                            (is_, diff_place) = check_field(seller_account.strip(), collectionAccount.strip())
                            advice += '【请核查收款账户，第{}个不同】；'.format(diff_place)
                            question_list.append(url_index)
                    except Exception as e:
                        # print(e)
                        collectionAccount_flag = 1
                        question_list.append(url_index)
                except Exception as e:
                    # advice = advice + "第{}张影像中机器人无法获取销售方开户行及账号值；".format(url_index)
                    seller_bank_account_flag = 1
                    question_list.append(url_index)

            if type == "1001":  # 增值税专用发票
                invoice_exists = True
                # 税额
                tax = item['details']['tax']
                if "," in tax:
                    tax = tax.replace(",", "")
                if tax == '':
                    tax = 0
                taxSum = taxSum + float(tax)  # 累加专项发票

            if type == "1044":  # 机票或者行程单
                invoice_exists = True
                # 仓位
                flights = item['details']['flights']
                for omg in flights:
                    seat = omg.get('seat', '')
                    if seat == '':
                        advice += '【第%s张影像，机票未识别到座位等级】' % url_index
                        flight_flag = 2
                    elif "F" in seat or "A" in seat or "C" in seat or "D" in seat:
                        # advice = advice + "第{}张影像中机票或行程单超仓；".format(url_index)
                        flight_flag = 1
                        question_list.append(url_index)
                    else:
                        # advice = advice + "第{}张影像中机票或行程单正常；".format(url_index)
                        flight_flag = 0

            if type == "1042":  # 火车票、高铁票
                invoice_exists = True
                # 仓位
                toSeat = item['details'].get('seat', '')
                if toSeat == '':
                    advice += '【第%s张影像，火车票或高铁票未识别到座位等级】' % url_index
                    seat_flag = 2
                elif "一等座" in toSeat or "商务座" in toSeat:
                    # advice = advice + "第{}张影像中火车票或高铁票超仓；".format(url_index)
                    seat_flag = 1
                    question_list.append(url_index)
                else:
                    seat_flag = 0
                    # advice = advice + "第{}张影像中火车票或高铁票正常；".format(url_index)

                    # if type != "1001" or type != "1002" or type != "1003" or type != "1004":
            #     advice = advice + "该单没有发票；"

    except Exception as e:
        print(e)
        print(traceback.format_exc())
        advice += "第{}张影像调用ocr接口失败；".format(url_index)

# HLT修改后逻辑
if bearing_flag == 0:
    advice += '购买方名称与费用承担公司一致；'
# elif bearing_flag == 1:
#     advice += '【购买方名称与费用承担公司不一致】；'
if taxnumber_flag == 0:
    advice += '税号一致；'
# elif taxnumber_flag == 1:
#     advice += '【税号不一致】；'
if seal_flag == 0:
    advice += '发票有盖章；'
elif seal_flag == 1:
    advice += '【发票没有有盖章】；'
if formname_flag == 0:
    advice += '发票有发票联；'
elif formname_flag == 1:
    advice += '【发票没有发票联】；'
if receivebank_flag == 0:
    advice += '收款银行与发票影像中的销售方开户行一致；'
# elif receivebank_flag == 1:
#     advice += '【收款银行与发票影像中的销售方开户行不一致】；'
if seller_flag == 0:
    advice += '收款人与发票中的销售方名称一致；'
# elif seller_flag == 1:
#     advice += '【收款人与发票中的销售方名称不一致】；'
if collectionAccount_flag == 0:
    advice += '收款账户与发票影像中的销售方账号一致；'
# elif collectionAccount_flag == 1:
#     advice += '【收款账户与发票影像中的销售方账号不一致】；'
if flight_flag == 0:
    advice += '机票或行程单正常；'
elif flight_flag == 1:
    advice += '【机票或行程单超仓】；'
if seat_flag == 0:
    advice += '火车票或高铁票正常；'
elif seat_flag == 1:
    advice += '【火车票或高铁票超仓】；'
if seller_bank_account_flag == 1:
    advice += '【接口未识别到销售方开户行及账号的值】；'

# 2021/1/15 ------------------------------
if credit_union_flag:
    advice += '【影像附件中发票含抵扣联】；'
if invoice_time_error_flag:
    advice += '【发票不为本年度发票】；'

# web:
finalPay = finalPay.replace(',', '')
final_pay = round(float(finalPay), 2)
moneyFee = moneyFee.replace(',', '')
finalPayTax = moneyFee
final_pay_tax = round(float(finalPayTax), 2)

## 只影响对外
if invoice_exists:
    if '对外付款单' in data['data'].get('current_title'):
        # 增加 由预付款单推出对外付款单审核要点
        subsist_no = data['data']['baseInfo'].get('预付款单号')
        pay_amount = data['data']['baseInfo'].get('付款金额')
        if float(pay_amount) == 0.0:
            if subsist_no not in ['', ' ', '...']:
                amount = data['data']['baseInfo'].get('合计应付金额')
                if amount:
                    final_pay = float(amount.replace(',', ''))

    if round(pay_sum, 2) != round(final_pay, 2):
        delta = abs(round(final_pay - pay_sum, 2))
        advice += '【请核查总金额, 差额%s】；' % str(delta)
    else:
        if '对外付款单' in data['data'].get('current_title'):
            advice += '总金额一致；'
    if round(taxSum, 2) != round(final_pay_tax, 2):
        advice += '【请核查税额，OCR显示{}】；'.format(str(round(taxSum, 2)))
        delta = abs(round(final_pay_tax - taxSum, 2))
        advice += '【请核查税额, 差额%s】；' % str(delta)
# 2021/1/15 ------------------------------

question_list = list(set(question_list))
if "不一致" in advice or "没有" in advice or "超仓" in advice:
    if len(question_list) == 0:
        pass
    else:
        a = ""
        for i in question_list:
            i = str(i)
            a += i + ','
        a = a[:-1]
        advice += '【请检查第{}张影像】；'.format(a)

# if advice_first == advice:
#     advice = advice + '【没有检测到发票】；'
listIata = totalSum

# 原有逻辑
# if "发票有盖章" in advice or "发票没有盖章" in advice:
#     advice = advice.replace("没有检测到发票；", "")
# else:
#     advice = advice + "没有检测到发票；"
#
# if "没有发票联；" in advice:
#     advice += "没有检测到发票；"


# 核对明细中的税额与专票增值税额
try:
    moneyFee = float(moneyFee)
except:
    moneyFee = 0
try:
    feeSum = float(feeSum)
except:
    feeSum = 0

try:
    if round(float(taxSum), 2) == round(float(moneyFee), 2):
        print("专票税额正常")
    elif round(float(taxSum), 2) == 0:
        print("专票税额正常")
    else:
        advice = advice + "【专票税额待核查】；"
except:
    pass

# 核对明细中的应付金额与所有发票的价税合计
try:
    if round(float(feeSum), 2) <= round(float(totalSum), 2):
        # float(totalSum)
        print("应付金额正常")
    elif round(float(totalSum), 2) == 0:
        print("应付金额正常")
    else:
        advice = advice + "【应付金额待核查】；"
except:
    pass

print(advice)

# except Exception as e:
#     print(e)
#     advice = advice + "没有检测到发票；"
# advice = advice.split("；")
# advice = list(set(advice))
# setAdvice = "；".join(advice)


if "没有检测到发票；" not in advice:
    taxSum = round(float(taxSum), 2)
    totalSum = round(float(totalSum) * 100) / 100.0
    if invoice_exists:
        advice = advice + "发票金额：" + str(totalSum) + ",专票税额：" + str(taxSum)
else:
    advice = advice.replace("没有检测到发票；", "")
    advice = advice.replace("【没有检测到发票】；", '')

if not invoice_exists:
    advice += "【没有检测到发票】；"
# 临时逻辑
if payee in ['中交（天津）疏浚工程有限公司', '中交天津航道局有限公司', '上海振华重工（集团）股份有限公司', '中交第一航务工程勘察设计院有限公司', '中交水运规划设计院有限公司', '中交一航局第一工程有限公司',
             '中交第一航务工程局有限公司'] or "中交" in payee:
    advice += "【暂停付款，打回本地】;"

data['data']['verifyResult'] = advice

approveopinions = data['data'].get('verifyResult', '')

# 摘要
abstract_list = baseInfo.get('abstract_list')
# 费用说明
fee_desc = baseInfo.get('fee_desc')
zip_abs_fee = zip(abstract_list, fee_desc)
new_abstract_list = list(map(lambda x: len(x[0].strip()) > 0 and x[0].strip() or x[1].strip(), zip_abs_fee))

# 摘要判断是否超过12个字，并且判断是否关联预付款单
abstract_length_error_list = list(map(lambda x: len(x) > 12, new_abstract_list))
abstract_relation_list = list(map(lambda x: x.find('结转') > -1 or x.find('核销') > -1, new_abstract_list))

if True in abstract_length_error_list:
    approveopinions += '【摘要字数大于十二字】；'
if True in abstract_relation_list:
    topay_no = baseInfo.get('预付款单号')
    if topay_no == '...' or '':
        approveopinions += '【单据未关联预付款单】；'

approveopinions = re.sub(r'[；，,]{2,}', '；', approveopinions)
approveopinions = re.sub(r'(.*?)【(审批流程无误；)】(.*)', r'\1\2\3', approveopinions)

p = re.compile('【(.*?)】')
L = p.findall(approveopinions)

for item in L:
    newItem = item.replace('；', '，').strip('，')
    approveopinions = approveopinions.replace(item, newItem)

approveopinions = approveopinions.replace('找不到封面页；请核实；', '【找不到封面页，请核实】；')
if approveopinions.find('机器人审批意见：') < 0:
    approveopinions = '机器人审批意见：' + approveopinions

pattern1 = re.compile(r'([^【】]*)(【[^【】]*?】)([^【】]*)')
pattern2 = re.compile(r'(机器人审批意见：)(.*)')
pattern3 = re.compile(r'；{2,}')

L = []
for i in pattern1.finditer(approveopinions):
    L.append(i.group(2))

approveopinions = pattern1.sub(r'\1\3', approveopinions)
approveopinions = pattern2.sub(r'\1%s；\2' % ''.join(L), approveopinions)
approveopinions = pattern3.sub('；', approveopinions).strip(',').strip('，')
approveopinions = approveopinions.replace('【提示：机器人试运行，；】', '')
data['data']['verifyResult'] = approveopinions


def record_runtime_data(runtime_data, name):
    with open(os.path.join(invoice_dir, name) + '.json', 'w', encoding='utf-8') as f:
        json.dump(runtime_data, f)


now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
with open(os.path.join(log_dir, 'OCR记录.txt'), 'a', encoding='gbk') as f:
    f.write('单据编号: ' + No + '\n')
    f.write('\n'.join(url_dict_new) + '\n')
    f.write(advice + '\n')
    record_runtime_data(log_list, 'ocr')
    record_runtime_data(data, '审批流')
    record = ['[', now, ']', '图片数量%s' % len(url_dict_new), '处理完成\n']
    f.write(''.join(record))
flow = json.dumps(data)
