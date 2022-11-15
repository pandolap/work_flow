#!/usr/bin/env python
# -*- coding: UTF-8 -*-


import json
import os
from datetime import datetime

import openpyxl
import requests

with open(r"") as f:
    flow = f.read()

data = json.loads(flow)

url = data['data']['imgN'][0]
local_company = data['data']['baseInfo']['公司']
a = ''
b = ''

log_dir = data['config'].get('log_dir')
No = data['data']['baseInfo']['单据编号']
invoice_dir = os.path.join(os.path.join(log_dir, 'ocr与审批流运行数据'), No)
if not os.path.exists(invoice_dir):
    os.makedirs(invoice_dir)


# local_company = "长海县广鹿码头建设管理有限公司"
def download_jpg(url):
    r = requests.get(url)
    now = datetime.now().strftime('%Y%m%d_%H%M%S')
    img_file = os.path.join(invoice_dir, now + '.jpg')
    with open(img_file, "wb") as f:
        f.write(r.content)
    return img_file


def get_ocr_text(file_path):
    request_url = "http://trial.web.glority.cn:8000/classify?type=20"
    # 二进制方式打开图片文件
    image_file = open(file_path, 'rb')
    img = {"image_file": image_file}
    response = requests.post(request_url, files=img)
    return response


# Todo url是影像系统里第一张图片的链接
f = download_jpg(url)

# 二进制方式打开图片文件
# f =r"C:\RPA\辽港表格范例--V1.0\1.jpg"

wb = openpyxl.load_workbook(r'C:\RPA\辽港表格范例--V1.0\辽港公司名单.xlsx')
ws = wb.get_sheet_by_name('辽港集团')
max_rows = ws.max_row
company_list = []
# 增加多次重试
ocr_flag = 0
tmp_response = get_ocr_text(f)
text = tmp_response.json()
while ocr_flag < 5:
    if text.get('response'):
        break
    else:
        ocr_flag += 1
        tmp_response = get_ocr_text(f)
        text = tmp_response.json()
# 将封面页ocr数据保存
cover_ocr_file = os.path.join(invoice_dir, '封面页OCR识别数据.json')
with open(cover_ocr_file, "w", encoding='utf-8') as f:
    json.dump(text, f)
if text["response"] and isinstance(text["response"], list) and isinstance(text["response"][0], dict) and text["response"][0]:
    words_list = text["response"][0]["text2"]
    for irow in range(max_rows):
        irow += 1
        company_list.append(ws.cell(irow, 2).value)
    for irow in range(max_rows):
        irow += 1
        sheet_company = ws.cell(irow, 2).value
        # Todo local_company 是前面取到的公司名称；
        if local_company == sheet_company:
            discrip = ws.cell(irow, 3).value
            if discrip is None:
                discrip = ""
            if "招商局港口" in words_list[:20]:
                b = "【封面打印有误；请检查；】"
                break
            if "股份" in discrip and "股份" in words_list[:20]:
                b = "打印封面正确；"
                break
            elif '太平湾' in discrip and '太平湾' in words_list[:20]:
                b = "打印封面正确；"
                break
            elif ("股份" not in discrip and '太平湾' not in discrip) and "集团" in words_list[:20]:
                b = "打印封面正确；"
                break
if b == "":
    b = "【封面打印有误；请检查；】"
data['data']['verifyResult'] = data['data']['verifyResult'] + a + b
flow = json.dumps(data)
